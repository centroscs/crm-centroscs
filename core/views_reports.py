from __future__ import annotations

import os
from datetime import datetime, date, time
from typing import Optional, List

from django.contrib.auth.decorators import login_required
from django.http import HttpRequest, HttpResponse
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.utils.dateparse import parse_date

from .models import Agent, Appointment, TodoItem, Property

# =========================
# ReportLab (PDF)
# =========================
from xml.sax.saxutils import escape as xml_escape

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image as RLImage,
)

# =========================
# Excel (XLSX)
# =========================
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# Helpers permessi
# ============================================================

def _is_admin_user(user) -> bool:
    return bool(user and (user.is_staff or user.is_superuser))


def _current_agent_for_request(request: HttpRequest) -> Optional[Agent]:
    try:
        return request.user.agent
    except Exception:
        return None


# ============================================================
# Helpers date/format
# ============================================================

def _parse_ymd(s: Optional[str]) -> Optional[date]:
    if not s:
        return None
    return parse_date(s.strip())


def _dt_start_of_day(d: date) -> datetime:
    tz = timezone.get_current_timezone()
    return timezone.make_aware(datetime.combine(d, time.min), tz)


def _dt_end_of_day(d: date) -> datetime:
    tz = timezone.get_current_timezone()
    return timezone.make_aware(datetime.combine(d, time.max), tz)


def _fmt_dt(dt: Optional[datetime]) -> str:
    if not dt:
        return "—"
    return timezone.localtime(dt).strftime("%d/%m/%Y %H:%M")


def _fmt_date(d: Optional[date]) -> str:
    if not d:
        return "—"
    return d.strftime("%d/%m/%Y")


def _now_str() -> str:
    return timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")


def _is_xlsx(request: HttpRequest) -> bool:
    return (request.GET.get("format") or "").strip().lower() == "xlsx"


# ============================================================
# ReportLab helpers
# ============================================================

def _pdf_response(filename: str) -> HttpResponse:
    resp = HttpResponse(content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    return resp


def _doc(resp: HttpResponse, title: str) -> SimpleDocTemplate:
    return SimpleDocTemplate(
        resp,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=title,
        author="CRM",
    )


def _footer(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillGray(0.35)
    canvas.drawString(12 * mm, 8 * mm, f"Generato: {_now_str()}")
    canvas.drawRightString(doc.pagesize[0] - 12 * mm, 8 * mm, f"Pagina {doc.page}")
    canvas.restoreState()


def _table_style() -> TableStyle:
    return TableStyle(
        [
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#efefef")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#c9c9c9")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fafafa")]),
        ]
    )


def _header_story(story, title: str, subtitle: str):
    styles = getSampleStyleSheet()
    story.append(Paragraph(title, styles["Heading2"]))
    story.append(Paragraph(subtitle, styles["Normal"]))
    story.append(Spacer(1, 6 * mm))


def _p(text: str, style: ParagraphStyle) -> Paragraph:
    safe = xml_escape(str(text)).replace("\n", "<br/>")
    return Paragraph(safe, style)


# ============================================================
# XLSX helpers
# ============================================================

def _xlsx_response(filename: str) -> HttpResponse:
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _xlsx_autofit(ws):
    # auto-width semplice (basato su lunghezze testo)
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 60)


def _xlsx_write_table(ws, headers: List[str], rows: List[List]):
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = bold

    for r in rows:
        ws.append(r)

    # wrap su tutto
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap

    _xlsx_autofit(ws)


# ============================================================
# INDEX REPORT
# ============================================================

@login_required
def reports_index(request: HttpRequest) -> HttpResponse:
    agents = Agent.objects.all().order_by("name") if _is_admin_user(request.user) else []
    return render(request, "core/reports_index.html", {"agents": agents})


# ============================================================
# REPORT APPUNTAMENTI (PDF o XLSX)
# filtri: from,to
# ============================================================

@login_required
def report_appointments_pdf(request: HttpRequest, agent_id: Optional[int] = None) -> HttpResponse:
    d_from = _parse_ymd(request.GET.get("from"))
    d_to = _parse_ymd(request.GET.get("to"))

    qs = Appointment.objects.select_related("agent", "contact", "property").all()

    agent_obj: Optional[Agent] = None
    if _is_admin_user(request.user):
        if agent_id is not None:
            agent_obj = get_object_or_404(Agent, pk=agent_id)
            qs = qs.filter(agent=agent_obj)
    else:
        agent_obj = _current_agent_for_request(request)
        if agent_obj:
            qs = qs.filter(agent=agent_obj)
        else:
            qs = qs.none()

    if d_from:
        qs = qs.filter(start__gte=_dt_start_of_day(d_from))
    if d_to:
        qs = qs.filter(start__lte=_dt_end_of_day(d_to))

    qs = qs.order_by("start")

    # ===== XLSX =====
    if _is_xlsx(request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Appuntamenti"

        headers = ["Data/Ora", "Titolo", "Agente", "Luogo", "Contatto", "Immobile"]
        rows = []
        for a in qs:
            rows.append([
                _fmt_dt(getattr(a, "start", None)),
                getattr(a, "title", "") or "(senza titolo)",
                getattr(getattr(a, "agent", None), "name", "") or "—",
                getattr(a, "location", "") or "—",
                getattr(getattr(a, "contact", None), "full_name", "") or "—",
                getattr(getattr(a, "property", None), "code", "") or "—",
            ])

        _xlsx_write_table(ws, headers, rows)

        resp = _xlsx_response("report_appuntamenti.xlsx" if not agent_obj else f"report_appuntamenti_{agent_obj.id}.xlsx")
        wb.save(resp)
        return resp

    # ===== PDF =====
    filename = "report_appuntamenti.pdf" if not agent_obj else f"report_appuntamenti_{agent_obj.id}.pdf"
    resp = _pdf_response(filename)
    doc = _doc(resp, "Report Appuntamenti")

    story: List = []
    subtitle = f"Generato: {_now_str()}"
    if agent_obj:
        subtitle += f" — Agente: {getattr(agent_obj, 'name', '')}"
    if d_from or d_to:
        subtitle += f" — Periodo: {_fmt_date(d_from)} → {_fmt_date(d_to)}"
    _header_story(story, "Report Appuntamenti", subtitle)

    data = [["Data/Ora", "Titolo", "Agente", "Luogo", "Contatto", "Immobile"]]
    for a in qs:
        data.append(
            [
                _fmt_dt(getattr(a, "start", None)),
                getattr(a, "title", "") or "(senza titolo)",
                getattr(getattr(a, "agent", None), "name", "") or "—",
                getattr(a, "location", "") or "—",
                getattr(getattr(a, "contact", None), "full_name", "") or "—",
                getattr(getattr(a, "property", None), "code", "") or "—",
            ]
        )

    if len(data) == 1:
        story.append(Paragraph("Nessun appuntamento trovato con i filtri selezionati.", getSampleStyleSheet()["Normal"]))
    else:
        tbl = Table(
            data,
            colWidths=[40 * mm, 55 * mm, 30 * mm, 55 * mm, 45 * mm, 25 * mm],
            repeatRows=1,
        )
        tbl.setStyle(_table_style())
        story.append(tbl)

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return resp


# ============================================================
# REPORT TODO (PDF o XLSX)
# filtri: status=open|done|all, from,to
# ============================================================

@login_required
def report_todos_pdf(request: HttpRequest, agent_id: Optional[int] = None) -> HttpResponse:
    status = (request.GET.get("status") or "open").strip().lower()
    if status not in ("open", "done", "all"):
        status = "open"

    d_from = _parse_ymd(request.GET.get("from"))
    d_to = _parse_ymd(request.GET.get("to"))

    qs = TodoItem.objects.select_related("agent").all()

    agent_obj: Optional[Agent] = None
    if _is_admin_user(request.user):
        if agent_id is not None:
            agent_obj = get_object_or_404(Agent, pk=agent_id)
            qs = qs.filter(agent=agent_obj)
    else:
        agent_obj = _current_agent_for_request(request)
        if agent_obj:
            qs = qs.filter(agent=agent_obj)
        else:
            qs = qs.none()

    if status == "open":
        qs = qs.filter(is_done=False)
    elif status == "done":
        qs = qs.filter(is_done=True)

    if d_from:
        qs = qs.filter(due_at__gte=_dt_start_of_day(d_from))
    if d_to:
        qs = qs.filter(due_at__lte=_dt_end_of_day(d_to))

    qs = qs.order_by("is_done", "due_at", "id")

    # ===== XLSX =====
    if _is_xlsx(request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Todo"

        headers = ["Scadenza", "Todo", "Agente", "Stato"]
        rows = []
        for t in qs:
            rows.append([
                _fmt_dt(getattr(t, "due_at", None)),
                getattr(t, "title", "") or "(senza titolo)",
                getattr(getattr(t, "agent", None), "name", "") or "—",
                "CHIUSA" if getattr(t, "is_done", False) else "APERTA",
            ])

        _xlsx_write_table(ws, headers, rows)

        resp = _xlsx_response("report_todo.xlsx" if not agent_obj else f"report_todo_{agent_obj.id}.xlsx")
        wb.save(resp)
        return resp

    # ===== PDF =====
    filename = "report_todo.pdf" if not agent_obj else f"report_todo_{agent_obj.id}.pdf"
    resp = _pdf_response(filename)
    doc = _doc(resp, "Report Todo")

    styles = getSampleStyleSheet()
    story: List = []

    subtitle = f"Generato: {_now_str()} — Stato: {status}"
    if agent_obj:
        subtitle += f" — Agente: {getattr(agent_obj, 'name', '')}"
    if d_from or d_to:
        subtitle += f" — Periodo: {_fmt_date(d_from)} → {_fmt_date(d_to)}"
    _header_story(story, "Report Todo", subtitle)

    data = [["Scadenza", "Todo", "Agente", "Stato"]]
    for t in qs:
        data.append(
            [
                _fmt_dt(getattr(t, "due_at", None)),
                getattr(t, "title", "") or "(senza titolo)",
                getattr(getattr(t, "agent", None), "name", "") or "—",
                "CHIUSA" if getattr(t, "is_done", False) else "APERTA",
            ]
        )

    if len(data) == 1:
        story.append(Paragraph("Nessuna todo trovata con i filtri selezionati.", styles["Normal"]))
    else:
        tbl = Table(
            data,
            colWidths=[45 * mm, 120 * mm, 40 * mm, 25 * mm],
            repeatRows=1,
        )
        tbl.setStyle(_table_style())
        story.append(tbl)

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return resp


# ============================================================
# REPORT IMMOBILI (PDF o XLSX)
# filtri: city, description, price_min, price_max, has_photo
# ============================================================

def _property_thumb_path(p: Property) -> Optional[str]:
    img_obj = None
    try:
        if hasattr(p, "primary_image") and callable(getattr(p, "primary_image")):
            img_obj = p.primary_image()
    except Exception:
        img_obj = None

    if img_obj is None:
        try:
            imgs = getattr(p, "images", None)
            if imgs is not None:
                img_obj = imgs.order_by("-is_primary", "position", "id").first()
        except Exception:
            img_obj = None

    if not img_obj:
        return None

    try:
        f = getattr(img_obj, "image", None)
        if f and hasattr(f, "path"):
            path = f.path
            if path and os.path.exists(path):
                return path
    except Exception:
        return None
    return None


@login_required
def report_properties_pdf(request: HttpRequest, agent_id: Optional[int] = None) -> HttpResponse:
    city = (request.GET.get("city") or "").strip()
    desc_q = (request.GET.get("description") or "").strip()

    price_min_raw = (request.GET.get("price_min") or "").strip()
    price_max_raw = (request.GET.get("price_max") or "").strip()
    has_photo = (request.GET.get("has_photo") or "").strip().lower() in ("1", "true", "on", "yes")

    qs = Property.objects.all()

    agent_obj: Optional[Agent] = None
    if _is_admin_user(request.user):
        if agent_id is not None:
            agent_obj = get_object_or_404(Agent, pk=agent_id)
            if hasattr(Property, "owner_agent"):
                qs = qs.filter(owner_agent=agent_obj)
    else:
        agent_obj = _current_agent_for_request(request)
        if agent_obj and hasattr(Property, "owner_agent"):
            qs = qs.filter(owner_agent=agent_obj)
        elif not _is_admin_user(request.user):
            qs = qs.none()

    if city:
        try:
            qs = qs.filter(city__icontains=city)
        except Exception:
            pass

    if desc_q:
        try:
            qs = qs.filter(description__icontains=desc_q)
        except Exception:
            pass

    # prezzo min/max (se il campo price è numerico)
    if price_min_raw:
        try:
            qs = qs.filter(price__gte=price_min_raw)
        except Exception:
            pass

    if price_max_raw:
        try:
            qs = qs.filter(price__lte=price_max_raw)
        except Exception:
            pass

    # solo con foto
    if has_photo:
        # se esiste relazione images, filtriamo con JOIN
        try:
            qs = qs.filter(images__isnull=False).distinct()
        except Exception:
            pass

    try:
        qs = qs.order_by("code")
    except Exception:
        qs = qs.order_by("id")

    # ===== XLSX =====
    if _is_xlsx(request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Immobili"

        headers = ["Codice", "Città", "Indirizzo", "Prezzo", "Descrizione", "Ha foto?"]
        rows = []

        for p in qs:
            code = getattr(p, "code", "") or "—"
            cty = getattr(p, "city", "") or "—"
            addr = getattr(p, "address", "") or "—"
            price = getattr(p, "price", None)
            price_str = f"{price}" if price not in (None, "") else "—"
            desc = getattr(p, "description", "") or "—"
            has_img = "SI" if _property_thumb_path(p) else "NO"

            rows.append([code, cty, addr, price_str, desc, has_img])

        _xlsx_write_table(ws, headers, rows)

        resp = _xlsx_response("report_immobili.xlsx" if not agent_obj else f"report_immobili_{agent_obj.id}.xlsx")
        wb.save(resp)
        return resp

    # ===== PDF =====
    filename = "report_immobili.pdf" if not agent_obj else f"report_immobili_{agent_obj.id}.pdf"
    resp = _pdf_response(filename)
    doc = _doc(resp, "Report Immobili")

    styles = getSampleStyleSheet()
    desc_style = ParagraphStyle(
        "desc",
        parent=styles["Normal"],
        fontSize=8,
        leading=9,
    )

    story: List = []
    subtitle = f"Generato: {_now_str()}"
    if agent_obj:
        subtitle += f" — Agente: {getattr(agent_obj, 'name', '')}"
    if city:
        subtitle += f" — Città: {city}"
    if desc_q:
        subtitle += f" — Descrizione contiene: “{desc_q}”"
    if price_min_raw:
        subtitle += f" — Prezzo ≥ {price_min_raw}"
    if price_max_raw:
        subtitle += f" — Prezzo ≤ {price_max_raw}"
    if has_photo:
        subtitle += " — Solo con foto"
    _header_story(story, "Report Immobili", subtitle)

    data = [["Foto", "Codice", "Città", "Indirizzo", "Prezzo", "Descrizione"]]

    thumb_w = 22 * mm
    thumb_h = 16 * mm

    for p in qs:
        code = getattr(p, "code", "") or "—"
        cty = getattr(p, "city", "") or "—"
        addr = getattr(p, "address", "") or "—"
        price = getattr(p, "price", None)
        price_str = f"€ {price}" if price not in (None, "") else "—"
        desc = getattr(p, "description", "") or "—"

        img_path = _property_thumb_path(p)
        if img_path:
            try:
                img = RLImage(img_path, width=thumb_w, height=thumb_h)
            except Exception:
                img = _p("—", desc_style)
        else:
            img = _p("—", desc_style)

        data.append([
            img,
            code,
            cty,
            _p(addr, desc_style),
            price_str,
            _p(desc, desc_style),
        ])

    if len(data) == 1:
        story.append(Paragraph("Nessun immobile trovato con i filtri selezionati.", styles["Normal"]))
    else:
        tbl = Table(
            data,
            colWidths=[26 * mm, 22 * mm, 30 * mm, 60 * mm, 25 * mm, 95 * mm],
            repeatRows=1,
        )
        st = _table_style()
        st.add("VALIGN", (0, 1), (0, -1), "MIDDLE")
        st.add("ALIGN", (0, 1), (0, -1), "CENTER")
        st.add("VALIGN", (3, 1), (5, -1), "TOP")
        tbl.setStyle(st)
        story.append(tbl)

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return resp
